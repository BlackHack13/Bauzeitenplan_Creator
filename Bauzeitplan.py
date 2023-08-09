import tkinter
import tkinter.messagebox
import tkinter.filedialog
import customtkinter
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import os
import calendar
import datetime

PATH = os.path.dirname(os.path.realpath(__file__)) + "/IMG"
customtkinter.set_appearance_mode("System")  # Modes: "System", "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue", "green", "dark-blue" "sweetkind"
MONATE = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember")
VERSION = "Version: 3.0"
LAENGE_SPALTE = 41


class GUI(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        aufloesung_faktor = {
            (1280, 720): (7, 1000, 800),  # HD
            (1366, 768): (7, 1000, 800),
            (1440, 900): (7, 1000, 800),
            (1600, 900): (7, 1000, 800),
            (1920, 1080): (6, 900, 800),  # FullHD
            (1680, 1050): (6, 900, 800),
            (1400, 1050): (6, 900, 800),
            (2048, 1152): (6, 900, 800),
            (2560, 1440): (5.5, 800, 800),  # QHD
            (2560, 1600): (5.5, 800, 800),
            (3200, 1800): (5, 800, 800),
            (3840, 2160): (5, 800, 800)}  # UHD

        faktor_und_groesse = aufloesung_faktor.get((screen_width, screen_height), (6, 900, 800))
        WIDTH = faktor_und_groesse[1]
        HEIGHT = faktor_und_groesse[2]

        self.title("Bauzeitenplan Creator")
        self.iconbitmap(PATH + '/icon.ico')
        self.geometry(f"{WIDTH}x{HEIGHT}")
        self.protocol(self.close_app)

        # ============ frames ============
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.frame_links = customtkinter.CTkFrame(master=self, width=400, corner_radius=0)
        self.frame_links.grid(row=0, column=0, sticky="nswe")

        self.frame_mitte = customtkinter.CTkFrame(master=self, width=180, corner_radius=0)
        self.frame_mitte.grid(row=0, column=1, sticky="nswe", padx=20, pady=20)

        self.frame_rechts = customtkinter.CTkFrame(master=self, width=180, corner_radius=0)
        self.frame_rechts.grid(row=0, column=2, sticky="nswe")

        # ==== grid layout (1x11) L ===
        self.frame_links.grid_rowconfigure(0, minsize=10)  # empty row with minsize as spacing
        self.frame_links.grid_rowconfigure(5, weight=1)  # empty row as spacing
        self.frame_links.grid_rowconfigure(8, minsize=20)  # empty row with minsize as spacing
        self.frame_links.grid_rowconfigure(20, minsize=10)  # empty row with minsize as spacing

        # === grid layout (3x7) M ===
        self.frame_mitte.rowconfigure((1, 2, 3), weight=1)
        self.frame_mitte.rowconfigure(8, weight=10)
        self.frame_mitte.columnconfigure((0, 1), weight=1)
        self.frame_mitte.columnconfigure(2, weight=0)

        # === grid layout (1x11) R ===
        self.frame_rechts.rowconfigure(0, weight=1)
        self.frame_rechts.rowconfigure(8, weight=10)
        self.frame_rechts.columnconfigure(8, weight=1)
        self.frame_rechts.columnconfigure(2, weight=0)

        # ////// Links
        self.logo = customtkinter.CTkImage(light_image=Image.open(PATH + "/logo.png"), dark_image=Image.open(PATH + "/logo.png"), size=(140, 60))
        self.label_logo = customtkinter.CTkLabel(master=self.frame_links, image=self.logo, textvariable="", text="")
        self.label_logo.grid(row=0, column=0, padx=(5, 0), pady=(5, 0))

        self.label_darstellung = customtkinter.CTkLabel(master=self.frame_links, text="Darstellung:")
        self.label_darstellung.grid(row=9, column=0)
        self.darstellung = customtkinter.CTkOptionMenu(master=self.frame_links, values=["Light", "Dark", "System"],
                                                       command=self.change_appearance_mode)
        self.darstellung.grid(row=10, column=0, pady=10, padx=20)
        self.darstellung.set("System")

        # ////// Mitte
        self.bauherr_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Bauherr: ", width=300,
                                                    height=25, border_width=2, corner_radius=5)
        self.bauherr_entry.grid(row=1, column=0, columnspan=3)
        self.projektname_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Projektname: ",
                                                        width=300, height=25, border_width=2, corner_radius=5)
        self.projektname_entry.grid(row=2, column=0, columnspan=3)
        self.projektnummer_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Projektnummer: ",
                                                          width=300, height=25, border_width=2, corner_radius=5)
        self.projektnummer_entry.grid(row=3, column=0, columnspan=3)

        self.startjahr_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Startjahr: ",
                                                      width=120, height=25, border_width=2, corner_radius=5)
        self.startjahr_entry.grid(row=5, column=0, padx=0, pady=10)
        self.endjahr_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Endjahr: ", width=120,
                                                    height=25, border_width=2, corner_radius=5)
        self.endjahr_entry.grid(row=5, column=1, padx=0, pady=10)

        self.startmonat_entry = customtkinter.CTkOptionMenu(master=self.frame_mitte, values=MONATE)
        self.startmonat_entry.grid(row=6, column=0)
        self.endmonat_entry = customtkinter.CTkOptionMenu(master=self.frame_mitte, values=MONATE)
        self.endmonat_entry.grid(row=6, column=1)

        self.baukosten = customtkinter.CTkCheckBox(master=self.frame_mitte, text="Baukosten")
        self.baukosten.grid(row=7, column=0, pady=10)

        self.button_speichern = customtkinter.CTkButton(master=self.frame_mitte, text="Speichern", width=300, height=25,
                                                        border_width=2, corner_radius=5, command=self.save_file)
        self.button_speichern.grid(row=9, column=0, pady=70)
        self.button_offnen = customtkinter.CTkButton(master=self.frame_mitte, text="Öffnen in Excel", width=300,
                                                     height=25, border_width=2, corner_radius=5, command=self.open)
        self.button_offnen.grid(row=9, column=1, pady=70)

        # ////// Rechts
        self.txt_o_baukosten = customtkinter.CTkLabel(master=self.frame_rechts, text="Ohne Baukosten")
        self.txt_o_baukosten.grid(row=1, column=0, pady=0, padx=20)
        img_o_baukosten = Image.open(PATH + "/prev1.png")
        img_scaled_o_baukosten = img_o_baukosten.resize((img_o_baukosten.width // 6, img_o_baukosten.height // 6))
        self.bg_img_o_baukosten = ImageTk.PhotoImage(img_scaled_o_baukosten)
        self.img_o_baukosten = tkinter.Label(master=self.frame_rechts, image=self.bg_img_o_baukosten, borderwidth=0)
        self.img_o_baukosten.grid(row=2, column=0)

        self.txt_baukosten = customtkinter.CTkLabel(master=self.frame_rechts, text="Mit Baukosten")
        self.txt_baukosten.grid(row=6, column=0, pady=0, padx=20)
        img_baukosten = Image.open(PATH + "/prev2.png")
        img_scaled_baukosten = img_baukosten.resize((img_baukosten.width // 6, img_baukosten.height // 6))
        self.bg_img_baukosten = ImageTk.PhotoImage(img_scaled_baukosten)
        self.img_baukosten = tkinter.Label(master=self.frame_rechts, image=self.bg_img_baukosten, borderwidth=0)
        self.img_baukosten.grid(row=7, column=0)

        self.label_name = customtkinter.CTkLabel(master=self.frame_rechts, text="Justin").grid(row=9, column=0)
        self.label_version = customtkinter.CTkLabel(master=self.frame_rechts, text=VERSION).grid(row=10, column=0,
                                                                                                 pady=10)

    def jahre_sind_zahlen(self):
        number1 = self.startjahr_entry.get()
        number2 = self.endjahr_entry.get()
        if not number1.isdigit() or not number2.isdigit():
            tkinter.messagebox.showerror("Fehler", "Ungültige Eingabe. Bitte geben Sie eine gültiges Jahr ein.")
            return False
        if number2 < number1:
            tkinter.messagebox.showerror("Fehler", "Startjahr muss kleiner Endjahr sein")
            return False
        return True

    def save_file(self):
        if not self.jahre_sind_zahlen():
            return
        eingabe = self.get_current_entries()
        dateiname = eingabe.projektnummer + " Bauzeitplan"
        self.filepath = tkinter.filedialog.asksaveasfilename(
            initialfile=dateiname,
            defaultextension=".xlsx",
            filetypes=[("Excel Arbeitsmappe", ".xlsx")])
        workbook = Excel(eingabe).get_workbook()
        workbook.save(self.filepath)

    def open(self):
        os.startfile(self.filepath)

    def get_current_entries(self):
        eingabe = Eingabe()
        eingabe.bauherr = self.bauherr_entry.get()
        eingabe.projektname = self.projektname_entry.get()
        eingabe.projektnummer = self.projektnummer_entry.get()
        eingabe.startjahr = int(self.startjahr_entry.get())
        eingabe.endjahr = int(self.endjahr_entry.get())
        eingabe.startmonat = MONATE.index(self.startmonat_entry.get()) + 1
        eingabe.endmonat = MONATE.index(self.endmonat_entry.get()) + 1
        eingabe.baukosten = bool(self.baukosten.get())
        return eingabe

    def close_app(self):
        self.destroy()

    def change_appearance_mode(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)


class Eingabe:
    def __init__(self):
        self.bauherr = ""
        self.projektname = ""
        self.projektnummer = ""
        self.startjahr = -1
        self.endjahr = -1
        self.startmonat = -1
        self.endmonat = -1
        self.baukosten = False


class Excel:
    def __init__(self, eingabe):
        self.workbook = openpyxl.Workbook()
        worksheet = self.workbook.active
        worksheet.title = eingabe.projektnummer + " Bauzeitplan"

        # === Basics===
        worksheet['A1'] = "Bauherr:"
        worksheet['B1'] = eingabe.bauherr
        worksheet['A2'] = "Projektnummer:"
        worksheet['B2'] = eingabe.projektnummer
        worksheet['A3'] = "Projektname:"
        worksheet['B3'] = eingabe.projektname
        worksheet['A6'] = "Anforderungen:"

        if eingabe.baukosten:
            self._baukosten_ja()

        self.jahr_monat_kw_excel(eingabe)

        self._prettify(eingabe)

    def get_workbook(self):
        return self.workbook

    @staticmethod
    def _spalten(eingabe):
        jahre = [x for x in range(eingabe.startjahr, eingabe.endjahr + 1)]
        kombi = [(jahr, monat) for jahr in jahre for monat in range(1, 13)]
        kombi = kombi[eingabe.startmonat - 1: (eingabe.endjahr - eingabe.startjahr) * 12 + eingabe.endmonat]
        kombi_mit_kalenderwochen = []

        for jahr, monat in kombi:
            kalenderwochen = []
            for woche in calendar.monthcalendar(jahr, monat):
                if woche[0] != 0:  # only include the week if it belongs to the current year
                    kalenderwochen.append(datetime.date(jahr, monat, woche[0]).isocalendar()[1])
            for kalenderwoche in kalenderwochen:
                kombi_mit_kalenderwochen.append((jahr, monat, kalenderwoche))
        return kombi_mit_kalenderwochen

    def jahr_monat_kw_excel(self, eingabe):
        worksheet = self.workbook.worksheets[0]
        yy_mm_kw = self._spalten(eingabe)
        spaltennummer_start = 6
        if eingabe.baukosten:
            spaltennummer_start = 8
        last_column_month = None
        last_column_month_start = spaltennummer_start
        last_column_month_end = spaltennummer_start
        last_column_year = None
        last_column_year_start = spaltennummer_start
        last_column_year_end = spaltennummer_start

        for iii, (jahr, monat, kalenderwoche) in enumerate(yy_mm_kw, 1):
            spaltennummer = spaltennummer_start + iii - 1
            worksheet.cell(row=3, column=spaltennummer).value = jahr
            worksheet.cell(row=4, column=spaltennummer).value = MONATE[monat - 1]
            worksheet.cell(row=5, column=spaltennummer).value = kalenderwoche

            if monat == last_column_month:
                last_column_month_end = spaltennummer
            else:
                if last_column_month is not None:
                    worksheet.merge_cells(start_row=4, start_column=last_column_month_start, end_row=4,
                                          end_column=last_column_month_end)
                last_column_month = monat
                last_column_month_start = spaltennummer
                last_column_month_end = spaltennummer

            if jahr == last_column_year:
                last_column_year_end = spaltennummer
            else:
                if last_column_year is not None:
                    worksheet.merge_cells(start_row=3, start_column=last_column_year_start, end_row=3,
                                          end_column=last_column_year_end)
                last_column_year = jahr
                last_column_year_start = spaltennummer
                last_column_year_end = spaltennummer

        worksheet.merge_cells(start_row=4, start_column=last_column_month_start, end_row=4,
                              end_column=last_column_month_end)
        worksheet.merge_cells(start_row=3, start_column=last_column_year_start, end_row=3,
                              end_column=last_column_year_end)

    def _baukosten_ja(self):
        worksheet = self.workbook.worksheets[0]
        worksheet.insert_cols(5, 2)
        worksheet.merge_cells('G4:G5')
        worksheet.merge_cells('F4:F5')
        worksheet['F4'] = "Masse"
        worksheet['G4'] = "Baukosten"

    def _prettify(self, eingabe):
        worksheet = self.workbook.worksheets[0]

        # === Merge Cells ===
        for row in range(1, 4):
            worksheet.merge_cells("B" + str(row) + ":E" + str(row))
        worksheet.merge_cells('A4:E5')
        worksheet.merge_cells('A6:E6')

        # === Merge Cells und Zähl hoch===
        durchnum = 0
        for row in range(7, LAENGE_SPALTE):
            durchnum += 1
            worksheet.merge_cells("A" + str(row) + ":E" + str(row))
            worksheet["A" + str(row)] = str(durchnum) + "."

        # Für Titel
        if eingabe.baukosten:
            worksheet['H1'] = "Bauzeitplan"
            worksheet.merge_cells(start_row=1, start_column=8, end_row=2, end_column=worksheet.max_column)
        else:
            worksheet['F1'] = "Bauzeitplan"
            worksheet.merge_cells(start_row=1, start_column=6, end_row=2, end_column=worksheet.max_column)

        # === Schriftgröße ===
        header1 = Font(bold=True)
        worksheet['B1'].font = header1
        worksheet['B2'].font = header1
        worksheet['B3'].font = header1
        worksheet['A6'].font = header1
        header2 = Font(bold=True, size=18)
        if eingabe.baukosten:
            worksheet['H1'].font = header2
        worksheet['F1'].font = header2

        # === Zell Breite ===
        dims = {}
        for row in worksheet.rows:
            if eingabe.baukosten:
                for cell in row[7:]:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                worksheet.column_dimensions["F"].width = 8
                worksheet.column_dimensions["G"].width = 12
            else:
                for cell in row[5:]:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = 3
        worksheet.column_dimensions["A"].width = 15

        # === Mittig Ausrichten ===
        for row in range(1, worksheet.max_row + 1):
            for col in range(6, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center')

        # === Rahmen ===
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))

        # === Farbigemakierung ===
        for zeile in range(6, LAENGE_SPALTE):
            for spalte in range(6, worksheet.max_column + 1):
                if worksheet.cell(row=5, column=spalte).value is not None and worksheet.cell(row=5,
                                                                                             column=spalte).value >= 52:
                    worksheet.cell(row=zeile, column=spalte).fill = PatternFill(fgColor='808080', fill_type='solid')


def main():
    # check_resolution()
    gui = GUI()
    gui.mainloop()


if __name__ == "__main__":
    main()
