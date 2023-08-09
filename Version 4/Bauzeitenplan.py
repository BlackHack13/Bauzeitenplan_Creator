import tkinter
import tkinter.messagebox
import tkinter.filedialog
import customtkinter
from PIL import Image
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import os
import keyboard
import calendar
import datetime

PATH = os.path.dirname(os.path.realpath(__file__)) + "/IMG"
customtkinter.set_appearance_mode("System")  # Modes: "System", "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue", "green", "dark-blue"
MONATE = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember")
WOCHENTAGE = ("Mo", "Di", "Mi", "Do", "Fr", "Sa", "So")
VERSION = 4.0



class GUI(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        check_version_window = CheckVersion(self)
        check_version_window.check_for_new_version()
        if check_version_window.winfo_exists():
            check_version_window.lift()
            self.wait_window(check_version_window)

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        aufloesung_faktor = {
            (1280, 720): (7, 1000, 800),  # HD
            (1366, 768): (7, 1000, 800),
            (1440, 900): (7, 1000, 800),
            (1600, 900): (7, 1000, 800),
            (1920, 1080): (6, 900, 800),  # FullHD
            (1680, 1050): (6, 900, 800),
            (1400, 1050): (6, 900, 800),
            (2048, 1152): (6, 900, 800),
            (2560, 1440): (5.5, 800, 800),  # QHD
            (2560, 1600): (5.5, 800, 800),
            (3200, 1800): (5, 800, 800),
            (3840, 2160): (5, 800, 800)}  # UHD
        faktor_und_groesse = aufloesung_faktor.get((screen_width, screen_height), (6, 900, 800))
        WIDTH = faktor_und_groesse[1]
        HEIGHT = faktor_und_groesse[2]

        self.title("Bauzeitenplan Creator")
        self.iconbitmap(PATH + '/icon.ico')
        self.geometry(f"{WIDTH}x{HEIGHT}")
        self.windows = {}
        self.protocol("WM_DELETE_WINDOW", self.on_main_window_close)

        # create the frames
        self.frame_oben = customtkinter.CTkFrame(self)
        self.frame_links = customtkinter.CTkFrame(self, corner_radius=0)
        self.frame_mitte = customtkinter.CTkFrame(self)
        self.frame_rechts = customtkinter.CTkFrame(self)

        # grid the frames
        self.frame_oben.grid(row=0, column=0, columnspan=3, sticky="nsew")
        self.frame_links.grid(row=1, column=0, padx=(0, 20), sticky="nsew")
        self.frame_mitte.grid(row=1, column=1, padx=(0, 20), pady=(10, 0), sticky="nsew")
        self.frame_rechts.grid(row=1, column=2, pady=(10, 0), sticky="nsew")


        # ==== grid layout (1x1) oben ===
        self.frame_oben.grid_rowconfigure(1, weight=1)
        self.frame_oben.grid_columnconfigure(1, weight=1)

        # configure grid layout (2x3)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(3, weight=0)
        self.grid_rowconfigure(1, weight=1)

        # ==== grid layout (1x11) L ===
        self.frame_links.grid_rowconfigure(0, minsize=10)
        self.frame_links.grid_rowconfigure(5, weight=1)
        self.frame_links.grid_rowconfigure(8, minsize=20)
        self.frame_links.grid_rowconfigure(20, minsize=10)

        # === grid layout (3x7) M ===
        self.frame_mitte.rowconfigure((1, 2, 3), weight=1)
        self.frame_mitte.rowconfigure(8, weight=10)
        self.frame_mitte.columnconfigure((0, 1), weight=1)
        self.frame_mitte.columnconfigure(2, weight=0)

        # === grid layout (1x11) R ===
        self.frame_rechts.rowconfigure(0, weight=1)
        self.frame_rechts.rowconfigure(8, weight=10)
        self.frame_rechts.columnconfigure(8, weight=1)
        self.frame_rechts.columnconfigure(2, weight=0)


        # ////// Top
        self.logo = customtkinter.CTkImage(light_image=Image.open(PATH + "/logo.png"), dark_image=Image.open(PATH + "/logo.png"), size=(140, 60))
        self.label_logo = customtkinter.CTkLabel(master=self.frame_oben, image=self.logo, textvariable="", text="")
        self.label_logo.grid(row=0, column=0, padx=(5, 0), pady=(5, 0))

        for i, window_name in enumerate(["EXCEL_Bearbeiten", "HOW_TO"], start=1):
            frame_button = customtkinter.CTkButton(self.frame_oben, text=f"Open {window_name}", command=lambda name=window_name: self.open_unique_window(name))
            frame_button.grid(row=0, column=i)

        # ////// Links
        self.label_darstellung = customtkinter.CTkLabel(self.frame_links, text="Darstellung:", anchor="w")
        self.label_darstellung.grid(row=6, column=0)
        self.darstellung_menu = customtkinter.CTkOptionMenu(self.frame_links, values=["Light", "Dark", "System"], command=self.change_appearance_mode)
        self.darstellung_menu.grid(row=7, column=0, pady=(0, 20))
        self.darstellung_menu.set("System")

        self.scaling_label = customtkinter.CTkLabel(self.frame_links, text="UI Skalierung:", anchor="w")
        self.scaling_label.grid(row=8, column=0)
        self.scaling_menu = customtkinter.CTkOptionMenu(self.frame_links, values=["50%", "75%", "100%", "125%", "150%", "200%"], command=self.change_scaling)
        self.scaling_menu.grid(row=9, column=0, padx=20, pady=(0, 20))
        self.scaling_menu.set("100%")

        # ////// Mitte
        self.bauherr_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Bauherr: ", width=300, height=25, border_width=2, corner_radius=5)
        self.bauherr_entry.grid(row=1, column=0, columnspan=3, padx=(5, 5))
        self.projektname_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Projektname: ", width=300, height=25, border_width=2, corner_radius=5)
        self.projektname_entry.grid(row=2, column=0, columnspan=3, padx=(5, 5))
        self.projektnummer_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Projektnummer: ", width=300, height=25, border_width=2, corner_radius=5)
        self.projektnummer_entry.grid(row=3, column=0, columnspan=3, padx=(5, 5))

        self.startjahr_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Startjahr: ", width=120, height=25, border_width=2, corner_radius=5)
        self.startjahr_entry.grid(row=5, column=0, padx=(5, 5), pady=15)
        self.endjahr_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Endjahr: ", width=120, height=25, border_width=2, corner_radius=5)
        self.endjahr_entry.grid(row=5, column=1, padx=(5, 5), pady=15)

        self.startmonat_entry = customtkinter.CTkOptionMenu(master=self.frame_mitte, values=MONATE)
        self.startmonat_entry.grid(row=6, column=0, padx=(5, 1))
        self.endmonat_entry = customtkinter.CTkOptionMenu(master=self.frame_mitte, values=MONATE)
        self.endmonat_entry.grid(row=6, column=1, padx=(1, 5))

        self.wochentage = customtkinter.CTkCheckBox(master=self.frame_mitte, text="Wochentage")
        self.wochentage.grid(row=7, column=0, pady=10)

        self.baukosten = customtkinter.CTkCheckBox(master=self.frame_mitte, text="Baukosten")
        self.baukosten.grid(row=7, column=1, pady=10)

        self.button_speichern = customtkinter.CTkButton(master=self.frame_mitte, text="Speichern", width=300, height=35, corner_radius=5, command=self.save_file)
        self.button_speichern.grid(row=9, column=0, padx=(5, 1), pady=(0, 60))
        self.button_offnen = customtkinter.CTkButton(master=self.frame_mitte, text="Öffnen in Excel", width=300, height=35, corner_radius=5, command=self.open, state="disabled")
        self.button_offnen.grid(row=9, column=1, padx=(1, 5), pady=(0, 60))

        # ////// Rechts
        self.txt_o_baukosten = customtkinter.CTkLabel(master=self.frame_rechts, text="Ohne Baukosten")
        self.txt_o_baukosten.grid(row=1, column=0, padx=(5, 5))
        self.img_o_baukosten = customtkinter.CTkImage(light_image=Image.open(PATH + "/prev1.png"), dark_image=Image.open(PATH + "/prev1.png"), size=(320, 170))
        self.label_img_o_baukosten = customtkinter.CTkLabel(master=self.frame_rechts, image=self.img_o_baukosten, text="")
        self.label_img_o_baukosten.grid(row=2, column=0, padx=(5, 5), pady=(0, 20))

        self.txt_baukosten = customtkinter.CTkLabel(master=self.frame_rechts, text="Mit Baukosten")
        self.txt_baukosten.grid(row=6, column=0, padx=(5, 5))
        self.img_baukosten = customtkinter.CTkImage(light_image=Image.open(PATH + "/prev2.png"), dark_image=Image.open(PATH + "/prev2.png"), size=(320, 170))
        self.label_img_baukosten = customtkinter.CTkLabel(master=self.frame_rechts, image=self.img_baukosten, text="")
        self.label_img_baukosten.grid(row=7, column=0, padx=(5, 5))

        self.label_name = customtkinter.CTkLabel(master=self.frame_rechts, text="Justin", anchor="n").grid(row=9, column=0)
        self.label_version = customtkinter.CTkLabel(master=self.frame_rechts, text=("Version: " + str(VERSION)), anchor="n").grid(row=10, column=0, pady=(0, 15))


        self.open_excel_file = False
    def open(self):
        if self.open_excel_file:
            os.startfile(self.filepath)

    def save_file(self):
        eingabe = self.get_current_entries()
        if eingabe is None:
            return

        eingabe = self.get_current_entries()
        dateiname = eingabe.projektnummer + " Bauzeitenplan"
        self.filepath = tkinter.filedialog.asksaveasfilename(
            initialfile=dateiname,
            defaultextension=".xlsx",
            filetypes=[("Excel Arbeitsmappe", ".xlsx")])
        workbook = Excel(eingabe).get_workbook()
        workbook.save(self.filepath)
        self.open_excel_file = True
        self.button_offnen.configure(state="normal")

    def get_current_entries(self):
        eingabe = Eingabe()
        eingabe.bauherr = self.bauherr_entry.get()
        eingabe.projektname = self.projektname_entry.get()
        eingabe.projektnummer = self.projektnummer_entry.get()
        eingabe.wochentage = bool(self.wochentage.get())
        eingabe.baukosten = bool(self.baukosten.get())

        # Überprüfen, ob Startjahr und Endjahr gültige Zahlen sind und Startjahr kleiner gleich Endjahr ist
        number1 = self.startjahr_entry.get()
        number2 = self.endjahr_entry.get()
        if not number1.isdigit() or not number2.isdigit():
            tkinter.messagebox.showerror("Fehler", "Jahre müssen eine Zahl sein")
            return None
        if int(number1) < 2000 or int(number2) < 2000:
            tkinter.messagebox.showerror("Fehler", "Jahre müssen ab 2000 sein")
            return None
        if int(number2) < int(number1):
            tkinter.messagebox.showerror("Fehler", "Startjahr muss kleiner oder gleich dem Endjahr sein.")
            return None
        eingabe.startjahr = int(number1)
        eingabe.endjahr = int(number2)

        # Überprüfen, ob Startmonat kleiner als Endmonat ist, wenn Start- und Endjahr gleich sind
        eingabe.startmonat = MONATE.index(self.startmonat_entry.get()) + 1
        eingabe.endmonat = MONATE.index(self.endmonat_entry.get()) + 1
        if eingabe.startjahr == eingabe.endjahr and eingabe.startmonat > eingabe.endmonat:
            tkinter.messagebox.showerror("Fehler", "Startmonat muss kleiner Endmonat sein, wenn Start- und Endjahr gleich sind.")
            return None

        return eingabe

    # GUI Funktionen
    def change_appearance_mode(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def change_scaling(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)

    def open_unique_window(self, window_name):
        window = self.windows.get(window_name)
        if not window:
            window = self.create_window(window_name)
            self.windows[window_name] = window
        self.lift()  # das geöffnete Fenster in den Vordergrund bringen
        self.iconify()  # das Hauptfenster minimieren

    def create_window(self, window_name):
        mapping = {"EXCEL_Bearbeiten": EXCEL_Bearbeiten,
                   "HOW_TO": HOW_TO}

        WindowClass = mapping[window_name]
        window = WindowClass(self)
        window.protocol("WM_DELETE_WINDOW", lambda: self.on_child_window_close(window))
        return window

    def on_child_window_close(self, window):
        window.destroy()  # das geöffnete Fenster zerstören
        self.windows = {k: v for k, v in self.windows.items() if v != window}  # das Fenster aus der windows-Dict entfernen
        self.wm_deiconify()  # das Hauptfenster wiederherstellen

    def on_main_window_close(self):
        self.destroy()  # das Hauptfenster schließen

    def toggle_fullscreen(self, event=None):
        self.attributes('-fullscreen', not self.attributes('-fullscreen'))
        if self.attributes('-fullscreen'):
            self.state('zoomed')  # Fenster maximieren
            self.wm_attributes('-topmost', 1)  # Fenster immer im Vordergrund anzeigen
        else:
            self.state('normal')  # Fensterzustand auf normal zurücksetzen
            self.wm_attributes('-topmost', 0)  # Standardmäßige Verhaltensweise wiederherstellen


class CheckVersion(customtkinter.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry("430x60")
        self.title("Neue Version")
        self.iconbitmap(PATH + '/icon.ico')

        self.path_to_versions_file = os.path.dirname(os.path.realpath(__file__))+ "/version.txt"
        self.current_version = VERSION

    def check_for_new_version(self):
        with open(self.path_to_versions_file, "r") as f:
            file_content = f.read()
        latest_version = float(file_content.strip())

        if latest_version > self.current_version:
            customtkinter.CTkLabel(self, text="Eine neue Version ist verfügbar! Bitte aktualisiere die Anwendung.").pack()
            customtkinter.CTkButton(self, text="OK", command=self.destroy).pack()
            self.lift()
        else:
            self.destroy()


class EXCEL_Bearbeiten(customtkinter.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry("200x200")
        self.title("EXCEL Bearbeiten")
        self.iconbitmap(PATH + '/icon.ico')
        customtkinter.CTkLabel(self, text="EXCEL Bearbeiten").pack()

class HOW_TO(customtkinter.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry("500x500")
        self.title("HOW TO")
        customtkinter.CTkLabel(self, text="HOW TO").pack()


class Eingabe:
    def __init__(self):
        self.bauherr = ""
        self.projektname = ""
        self.projektnummer = ""
        self.startjahr = -1
        self.endjahr = -1
        self.startmonat = -1
        self.endmonat = -1
        self.baukosten = False
        self.wochentage = False

class Excel:
    def __init__(self, eingabe):
        self.workbook = openpyxl.Workbook()
        kombi_mit_kalenderwochen = self._spalten(eingabe)
        worksheet = self.workbook.active
        worksheet.title = eingabe.projektnummer + " Bauzeitenplan"

        # === Basics===
        worksheet['A1'] = "Bauherr:"
        worksheet['B1'] = eingabe.bauherr
        worksheet['A2'] = "Projektnummer:"
        worksheet['B2'] = eingabe.projektnummer
        worksheet['A3'] = "Projektname:"
        worksheet['B3'] = eingabe.projektname

        worksheet['A4'] = "Stand: " + datetime.datetime.now().date().strftime("%d.%m.%Y") + "\n(Bauzeit " + str(len(set((x[0], x[1]) for x in kombi_mit_kalenderwochen))) + " Monate)"
        worksheet['A4'].alignment = Alignment(wrap_text=True)

        worksheet['A6'] = "Anforderungen:"
        worksheet['C3'] = "Jahr"
        worksheet['C4'] = "Monat"
        worksheet['C5'] = "Woche"

        if eingabe.baukosten:
            self._baukosten_ja()

        self.yy_mm_kw_dd_excel(eingabe)
        self.prettify(eingabe)

    def prettify(self, eingabe):
        worksheet = self.workbook.worksheets[0]
        finetune = Finetune(eingabe, worksheet)
        methods = [finetune._merge_cells, finetune._count_up, finetune._title, finetune._font_size, finetune._cell_width, finetune._cell_text, finetune._cell_border, finetune._cell_hatch]
        for method in methods:
            method(worksheet)

    def get_workbook(self):
        return self.workbook

    @staticmethod
    def _spalten(eingabe):
        first_month = (eingabe.startjahr, eingabe.startmonat)
        last_month = (eingabe.endjahr, eingabe.endmonat)

        # Generate a list of all months between the start and end years
        jahre = [x for x in range(eingabe.startjahr, eingabe.endjahr + 1)]
        kombi = [(jahr, monat) for jahr in jahre for monat in range(1, 13)]
        kombi = [m for m in kombi if first_month <= m <= last_month]

        # Create a list of (year, month, calendar week) tuples for each month in kombi
        kombi_mit_kalenderwochen = []
        for jahr, monat in kombi:
            kalenderwochen = []
            for woche in calendar.monthcalendar(jahr, monat):
                woche = [x for x in woche if x != 0]
                kalenderwochen.append(datetime.date(jahr, monat, woche[0]).isocalendar()[1])
            for kalenderwoche in kalenderwochen:
                kombi_mit_kalenderwochen.append((jahr, monat, kalenderwoche))

        return kombi_mit_kalenderwochen

    def yy_mm_kw_dd_excel(self, eingabe):
        worksheet = self.workbook.worksheets[0]
        yy_mm_kw = self._spalten(eingabe)
        spaltennummer_start = 4
        if eingabe.baukosten:
            spaltennummer_start = 6
        last_column_week = None
        last_column_week_start = spaltennummer_start
        last_column_week_end = spaltennummer_start
        last_column_month = None
        last_column_month_start = spaltennummer_start
        last_column_month_end = spaltennummer_start
        last_column_year = None
        last_column_year_start = spaltennummer_start
        last_column_year_end = spaltennummer_start

        spaltennummer = spaltennummer_start - 1

        for jahr, monat, kalenderwoche in yy_mm_kw:

            spaltennummer += 1

            worksheet.cell(row=3, column=spaltennummer).value = jahr
            worksheet.cell(row=4, column=spaltennummer).value = MONATE[monat - 1]
            worksheet.cell(row=5, column=spaltennummer).value = kalenderwoche

            # Füge Wochentage hinzu
            if eingabe.wochentage:
                for tag in WOCHENTAGE:
                    spaltennummer += 1
                    worksheet.cell(row=6, column=spaltennummer).value = tag[:2].upper()

            if kalenderwoche == last_column_week:
                last_column_week_end = spaltennummer
            else:
                if last_column_week is not None:
                    worksheet.merge_cells(start_row=5, start_column=last_column_week_start, end_row=5, end_column=last_column_week_end)
                last_column_week = kalenderwoche
                last_column_week_start = spaltennummer
                last_column_week_end = spaltennummer

            if monat == last_column_month:
                last_column_month_end = spaltennummer
            else:
                if last_column_month is not None:
                    worksheet.merge_cells(start_row=4, start_column=last_column_month_start, end_row=4, end_column=last_column_month_end)
                last_column_month = monat
                last_column_month_start = spaltennummer
                last_column_month_end = spaltennummer

            if jahr == last_column_year:
                last_column_year_end = spaltennummer
            else:
                if last_column_year is not None:
                    worksheet.merge_cells(start_row=3, start_column=last_column_year_start, end_row=3, end_column=last_column_year_end)
                last_column_year = jahr
                last_column_year_start = spaltennummer
                last_column_year_end = spaltennummer

        worksheet.merge_cells(start_row=4, start_column=last_column_month_start, end_row=4, end_column=last_column_month_end)
        worksheet.merge_cells(start_row=3, start_column=last_column_year_start, end_row=3, end_column=last_column_year_end)
        Finetune.merge_calendar_cells(worksheet, spaltennummer_start)


    def _baukosten_ja(self):
        worksheet = self.workbook.worksheets[0]
        worksheet.insert_cols(4, 2)
        worksheet.merge_cells('C4:C5')
        worksheet.merge_cells('D4:D5')
        worksheet['C4'] = "Masse"
        worksheet['D4'] = "Baukosten"

        worksheet['C3'] = ""
        worksheet['E3'] = "Jahr"
        worksheet['E4'] = "Monat"
        worksheet['E5'] = "Woche"

class Finetune:
    LAENGE_SPALTE = 41

    def __init__(self, eingabe, workbook):
        self.eingabe = eingabe
        self.workbook = workbook

    def _merge_cells(self, worksheet):
        worksheet.merge_cells('A4:A5')
        worksheet.merge_cells('B4:B5')
        if self.eingabe.baukosten:
            worksheet.merge_cells('A6:B6')
        else:
            worksheet.merge_cells('A6:C6')

    def _count_up(self, worksheet):
        durchnum = 0
        for row in range(7, Finetune.LAENGE_SPALTE):
            durchnum += 1
            if self.eingabe.baukosten:
                worksheet.merge_cells("A" + str(row) + ":B" + str(row))
                worksheet["A" + str(row)] = str(durchnum) + "."
            else:
                worksheet.merge_cells("A" + str(row) + ":C" + str(row))
                worksheet["A" + str(row)] = str(durchnum) + "."

    def _title(self, worksheet):
        if self.eingabe.baukosten:
            worksheet['F1'] = "Bauzeitenplan"
            worksheet.merge_cells(start_row=1, start_column=6, end_row=2, end_column=worksheet.max_column)
        else:
            worksheet['D1'] = "Bauzeitenplan"
            worksheet.merge_cells(start_row=1, start_column=4, end_row=2, end_column=worksheet.max_column)

    def _font_size(self, worksheet):
        header1 = Font(bold=True)
        worksheet['B1'].font = header1
        worksheet['B2'].font = header1
        worksheet['B3'].font = header1
        worksheet['A6'].font = header1
        header2: Font = Font(bold=True, size=18)
        if self.eingabe.baukosten:
            worksheet['F1'].font = header2
        worksheet['D1'].font = header2

    def _cell_width(self, worksheet):
        dims = {}
        for row in worksheet.rows:
            if self.eingabe.baukosten:
                for cell in row[5:]:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                worksheet.column_dimensions["C"].width = 8
                worksheet.column_dimensions["D"].width = 12
            else:
                for cell in row[3:]:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = 3

        worksheet.column_dimensions["A"].width = 20
        worksheet.freeze_panes = 'C1'

    @staticmethod
    def merge_calendar_cells(worksheet, min_col):
        merged_ranges = set(worksheet.merged_cells.ranges.copy())
        for merged in merged_ranges:
            # Überprüfe, ob der zusammengeführte Zellbereich in der fünften Zeile beginnt und die Mindestspalte überschreitet
            if merged.min_row == 5 and merged.min_col > min_col:
                # Zusammenführung der Zellen
                for row in range(6, Finetune.LAENGE_SPALTE):
                    worksheet.merge_cells(start_row=row, start_column=merged.min_col, end_row=row, end_column=merged.max_col)
    def _cell_text(self, worksheet):    # Zentriere den Text in der Zelle
        for row in range(1, worksheet.max_row + 1):
            for col in range(4, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center')

    def _cell_border(self, worksheet):
        # Iteriere über die Zeilen und Spalten des Arbeitsblatts
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                # dünnen Rahmen für alle Zellen
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                # ellenwert in der fünften Zeile
                if worksheet.cell(row=5, column=col).value is not None and (worksheet.cell(row=5, column=col).value == 1):
                    # dicken Rahmen für die Zelle
                    cell.border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def _cell_hatch(self, worksheet):
        # Wochen, die formatiert werden sollen
        weeks_to_format = {1, 32, 33, 52, 53, 54}
        # Iteriere über die Zeilen und Spalten des Arbeitsblatts
        for row in range(6, Finetune.LAENGE_SPALTE):
            for col in range(4, worksheet.max_column + 1):
                # Wert der Zelle in der fünften Zeile der aktuellen Spalte
                cell_value = worksheet.cell(row=5, column=col).value
                # Überprüfe, ob der Zellenwert in den zu formatierenden Wochen enthalten ist
                if cell_value is not None and cell_value in weeks_to_format:
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = PatternFill(fgColor='808080', fill_type='solid')
                    # Formatierung für bestimmte Zeilen
                    if 20 <= row <= 26:
                        cell.value = "Bauferien"
                        cell.font = Font(size=16, color='FFFFFF')
                        cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

        # Verbinde die Zellen mit "Bauferien"
        self.merge_cells_with_value(worksheet, "Bauferien")

    @staticmethod
    def merge_cells_with_value(worksheet, value):
        merged_ranges = []
        merge_start = None
        merge_end = None
        # Iteriere über die Zeilen und Spalten des Arbeitsblatts
        for row in range(6, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=4)
            if cell.value == value:
                # Wenn die aktuelle Zelle den Wert "Bauferien" enthält
                if merge_start is None:
                    merge_start = cell.coordinate
                merge_end = cell.coordinate
            else:
                # Wenn der vorherige Bereich beendet wurde, füge ihn zur Liste der zusammengeführten Bereiche hinzu
                if merge_start is not None:
                    merged_ranges.append((merge_start, merge_end))
                    merge_start = None
                    merge_end = None

        # Führe die letzten zusammenhängenden Zellen hinzu, falls vorhanden
        if merge_start is not None and merge_end is not None:
            merged_ranges.append((merge_start, merge_end))

        # Verbinde die Zellen
        for merge_range in merged_ranges:
            worksheet.merge_cells(start_row=int(merge_range[0][1:]), start_column=4, end_row=int(merge_range[1][1:]), end_column=4)


class Keyboard_Input:
    def __init__(self, gui):
        self.gui = gui

    def keyboard_input(self, event):
        if event.event_type == keyboard.KEY_DOWN:
            if event.name == 'f11' or (keyboard.is_pressed('ctrl') and event.name == 'f'): # Strg+ F; F11 Vollbild
                self.gui.toggle_fullscreen()

            elif keyboard.is_pressed('ctrl') and event.name == 'q':  # Strg+ Q Fenster schließen
                self.gui.on_main_window_close()

            elif keyboard.is_pressed('ctrl') and event.name == 'o':  # Strg+ O Datei öffnen
                self.gui.open()

            elif keyboard.is_pressed('ctrl') and event.name == 's':  # Strg+ S Datei speichern
                self.gui.save_file()

    def start_keyboard_input(self):
        keyboard.on_press(self.keyboard_input)

def main():
    gui = GUI()
    keyboard_input = Keyboard_Input(gui)
    keyboard_input.start_keyboard_input()

    # Rest des Hauptprogramms
    gui.mainloop()

if __name__ == "__main__":
    main()


