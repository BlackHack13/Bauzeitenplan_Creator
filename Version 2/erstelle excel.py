import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import os

PATH = os.path.dirname(os.path.realpath(__file__))
länge_spalte = 57
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
medium_border = Border(left=Side(style='medium'),
                       right=Side(style='medium'),
                       top=Side(style='medium'),
                       bottom=Side(style='medium'))
thick_border = Border(left=Side(style='thick'),
                      right=Side(style='thick'),
                      top=Side(style='thick'))
wb = openpyxl.Workbook()
worksheet = wb.active
bottom = Side(style='thick')

class Excel:

    bauherr = "John"
    projektnummer = "19-0718"
    projektname = projektnummer + ", " + bauherr
    worksheet.title = projektnummer + " " + "Bauzeitplan"

    # === Basics===
    worksheet['A1'] = "Bauherr:"
    worksheet['B1'] = bauherr
    worksheet['A2'] = "Projektnummer:"
    worksheet['B2'] = projektnummer
    worksheet['A3'] = "Projektname:"
    worksheet['B3'] = projektname
    worksheet['A6'] = "Anforderungen"
    worksheet['F1'] = "Bauzeitplan"
    wb.save(PATH + '/Bauzeitplan.xlsx')

    def baukosten_ja():
        workbook = openpyxl.load_workbook(PATH + '/Bauzeitplan.xlsx')
        worksheet = workbook.worksheets[0]
        worksheet.insert_cols(5, 2)
        worksheet.merge_cells('G4:G5')
        worksheet.merge_cells('F4:F5')
        worksheet['F4'] = "Masse"
        worksheet['G4'] = "Baukosten"
        for baukosten_num in range(ord("F"), ord("H")):
            for bau_row_num in range(1, länge_spalte):
                worksheet[chr(baukosten_num) + str(bau_row_num)].border = thin_border
        workbook.save(PATH + '/Bauzeitplan.xlsx')


    def prettify():
        workbook = openpyxl.load_workbook(PATH + '/Bauzeitplan.xlsx')
        worksheet = workbook.worksheets[0]

        #=== Zell Breite ===
        for spalte_size in range(ord("A"), ord("Z")):
            worksheet.column_dimensions[chr(spalte_size)].auto_size = True

        #=== Merge Cells ===
        for row in range(1, 4):
            worksheet.merge_cells("B" + str(row) + ":E" + str(row))
        worksheet.merge_cells('A4:E5')

        worksheet.merge_cells('A6:E6')
        durchnum = 0
        for row in range(7, länge_spalte):
            durchnum += 1
            worksheet.merge_cells("A" + str(row) + ":E" + str(row))
            worksheet["A" + str(row)] = str(durchnum) + "."

        worksheet.merge_cells('F1:BE2')

        # === Umgrenzung ===
        for spalte_num in range(ord("A"), ord("F")):
            for row_num in range(1, länge_spalte):
                worksheet[chr(spalte_num) + str(row_num)].border = thin_border

        # === Text Ausrichtung ===
        for spalte_num in range(ord("A"), ord("Z")):
            for row_num in range(1, 7):
                worksheet[chr(spalte_num) + str(row_num)].alignment = Alignment(horizontal='center', vertical='center')

        worksheet['F1'].alignment = Alignment(horizontal='center', vertical='center')

        #=== SAVE ===
        workbook.save(PATH + '/Bauzeitplan.xlsx')


    baukosten_ja()
    prettify()

os.startfile(PATH + '/Bauzeitplan.xlsx')
if __name__ == "__main__":
    Excel()

